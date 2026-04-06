"""
Excel Report Generator
========================
Produces a dark-themed, styled .xlsx export of I/O log records.
"""

import os
from datetime import datetime

import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ── Style constants ────────────────────────────────────────────────────────
_BG_DARK    = PatternFill("solid", fgColor="050709")
_BG_PANEL   = PatternFill("solid", fgColor="0D1117")
_BG_ROW     = PatternFill("solid", fgColor="111820")
_BG_WARN    = PatternFill("solid", fgColor="2A0E00")

_F_HEADER   = Font(bold=True, color="00E5FF",  name="Calibri", size=11)
_F_NORMAL   = Font(color="C8D8E8",             name="Calibri", size=10)
_F_GREEN    = Font(color="00FF88",             name="Calibri", size=10)
_F_WARN     = Font(bold=True, color="FF6D00",  name="Calibri", size=10)
_F_MUTED    = Font(color="3A4A5A",             name="Calibri", size=10)

_BORDER = Border(
    left=Side(style="thin", color="1A2230"),
    right=Side(style="thin", color="1A2230"),
    top=Side(style="thin", color="1A2230"),
    bottom=Side(style="thin", color="1A2230"),
)
_CENTER = Alignment(horizontal="center", vertical="center")
_LEFT   = Alignment(horizontal="left",   vertical="center")

# Column definitions: (header label, width)
_COLUMNS = [
    ("Timestamp", 22), ("PLC", 16), ("Device", 10), ("Label", 22),
    ("Type", 8), ("Value", 10), ("Warning", 28), ("Status", 12),
]


def _write_title(ws, total: int, warn_count: int):
    """Write the merged title banner rows."""
    ws.merge_cells("A1:H1")
    c = ws["A1"]
    c.value     = "PLC I/O MONITOR — DATA LOG"
    c.font      = Font(bold=True, color="00E5FF", name="Calibri", size=14)
    c.fill      = _BG_DARK
    c.alignment = _CENTER
    ws.row_dimensions[1].height = 28

    ws.merge_cells("A2:H2")
    c = ws["A2"]
    c.value = (
        f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}  |  "
        f"Records: {total}  |  Warnings: {warn_count}"
    )
    c.font      = Font(color="5A7080", name="Calibri", size=9)
    c.fill      = _BG_DARK
    c.alignment = _CENTER
    ws.row_dimensions[2].height = 15


def _write_headers(ws):
    """Write column header row."""
    for col, (label, width) in enumerate(_COLUMNS, start=1):
        cell            = ws.cell(row=3, column=col, value=label)
        cell.font       = _F_HEADER
        cell.fill       = _BG_PANEL
        cell.alignment  = _CENTER
        cell.border     = _BORDER
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.row_dimensions[3].height = 20


def _write_data_rows(ws, records: list):
    """Write one row per log record."""
    for row_idx, rec in enumerate(records, start=4):
        has_warning = bool(rec.get("warning"))
        fill = _BG_WARN if has_warning else (_BG_ROW if row_idx % 2 == 0 else _BG_PANEL)

        value_str  = "ON" if (rec["type"] == "bit" and rec["value"]) else (
                     "OFF" if rec["type"] == "bit" else str(rec["value"]))
        status_str = "⚠ WARNING" if has_warning else "✓ OK"
        timestamp  = rec["timestamp"][:19].replace("T", " ")

        row_data = [
            timestamp, rec["plc_name"], rec["device"], rec["label"],
            rec["type"].upper(), value_str, rec.get("warning", ""), status_str,
        ]

        for col, value in enumerate(row_data, start=1):
            cell            = ws.cell(row=row_idx, column=col, value=value)
            cell.fill       = fill
            cell.border     = _BORDER
            cell.alignment  = _CENTER if col in (1, 3, 5, 6, 8) else _LEFT

            # Value column
            if col == 6:
                if rec["type"] == "bit":
                    cell.font = _F_GREEN if rec["value"] else _F_MUTED
                else:
                    cell.font = _F_WARN if has_warning else _F_GREEN
            # Warning column
            elif col == 7:
                cell.font = _F_WARN if has_warning else _F_MUTED
            # Status column
            elif col == 8:
                cell.font = Font(
                    color="FF6D00" if has_warning else "00FF88",
                    name="Calibri", size=10, bold=True,
                )
            else:
                cell.font = _F_NORMAL

        ws.row_dimensions[row_idx].height = 15


def _write_summary_sheet(wb, records: list):
    """Add a Summary worksheet."""
    ws = wb.create_sheet("Summary")
    ws.sheet_view.showGridLines = False

    ws["A1"].value     = "SUMMARY"
    ws["A1"].font      = Font(bold=True, color="00E5FF", name="Calibri", size=13)
    ws["A1"].fill      = _BG_DARK
    ws.row_dimensions[1].height = 24

    warn_count = sum(1 for r in records if r.get("warning"))
    summary_rows = [
        ("Total Records",   len(records)),
        ("Warning Records", warn_count),
        ("OK Records",      len(records) - warn_count),
        ("Unique Devices",  len(set(r["device"]   for r in records))),
        ("PLCs Logged",     len(set(r["plc_name"] for r in records))),
        ("Log Start",       records[0]["timestamp"][:19].replace("T", " ") if records else "N/A"),
        ("Log End",         records[-1]["timestamp"][:19].replace("T", " ") if records else "N/A"),
    ]

    for row_idx, (key, val) in enumerate(summary_rows, start=3):
        c1 = ws.cell(row=row_idx, column=1, value=key)
        c1.font = Font(color="5A7080", name="Calibri", size=10)
        c1.fill = _BG_PANEL

        c2 = ws.cell(row=row_idx, column=2, value=str(val))
        c2.font = Font(color="C8D8E8", name="Calibri", size=10)
        c2.fill = _BG_PANEL

        # Highlight warning count in orange
        if row_idx == 4 and int(str(val)) > 0:
            c2.font = Font(bold=True, color="FF6D00", name="Calibri", size=10)

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 25


def make_excel(records: list, filename: str, logs_dir: str = "logs") -> str:
    """
    Generate a styled Excel workbook from `records`.

    Args:
        records:  List of log record dicts from state.log_records.
        filename: The .xlsx filename (saved inside `logs_dir`).
        logs_dir: Directory to save the file in (default: 'logs').

    Returns:
        The full path to the saved file.
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "IO Log"
    ws.sheet_view.showGridLines = False

    warn_count = sum(1 for r in records if r.get("warning"))
    _write_title(ws, len(records), warn_count)
    _write_headers(ws)
    _write_data_rows(ws, records)
    _write_summary_sheet(wb, records)

    path = os.path.join(logs_dir, filename)
    wb.save(path)
    return path
