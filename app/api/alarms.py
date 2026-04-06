"""
API — Alarm History
=====================
Persists a rolling log of alarm events (raised / acknowledged / cleared).

Routes:
  GET  /api/alarms/history            → return alarm history list
  POST /api/alarms/acknowledge/<id>   → ack a specific alarm by id
  POST /api/alarms/acknowledge/all    → ack all unacknowledged alarms
  POST /api/alarms/clear              → wipe history
  GET  /api/alarms/export             → download Excel report
"""

import os
import uuid
from datetime import datetime
from flask import Blueprint, jsonify, request, send_from_directory
from app import state

bp = Blueprint("alarms", __name__, url_prefix="/api/alarms")

# In-memory alarm history (newest first)
alarm_history: list = []   # Each entry: {id, ts_raised, ts_acked, plc_name, device, message, severity, acked}

_LOGS_DIR = os.path.join(
    os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__)))), "logs"
)
os.makedirs(_LOGS_DIR, exist_ok=True)


def record_alarm(plc_id: int, plc_name: str, device: str, message: str, severity: str = "warning"):
    """Add an event to alarm history. Called when alarm is raised."""
    entry = {
        "id":        str(uuid.uuid4())[:8],
        "ts_raised": datetime.now().isoformat(timespec="seconds"),
        "ts_acked":  None,
        "plc_id":    plc_id,
        "plc_name":  plc_name,
        "device":    device,
        "message":   message,
        "severity":  severity,
        "acked":     False,
    }
    alarm_history.insert(0, entry)
    # Trim to 1000 events
    if len(alarm_history) > 1000:
        alarm_history.pop()
    return entry


@bp.get("/history")
def get_history():
    limit = int(request.args.get("limit", 200))
    sev   = request.args.get("severity", "")       # filter by severity
    acked = request.args.get("acked", "")          # "true"/"false"/""
    plc   = request.args.get("plc", "")

    items = alarm_history
    if sev:
        items = [x for x in items if x["severity"] == sev]
    if acked == "true":
        items = [x for x in items if x["acked"]]
    elif acked == "false":
        items = [x for x in items if not x["acked"]]
    if plc:
        items = [x for x in items if x["plc_name"] == plc]

    return jsonify({
        "success":      True,
        "total":        len(alarm_history),
        "unacked":      sum(1 for x in alarm_history if not x["acked"]),
        "alarms":       items[:limit],
    })


@bp.post("/acknowledge/all")
def ack_all():
    now = datetime.now().isoformat(timespec="seconds")
    count = 0
    for a in alarm_history:
        if not a["acked"]:
            a["acked"]    = True
            a["ts_acked"] = now
            count += 1
    return jsonify({"success": True, "acknowledged": count})


@bp.post("/acknowledge/<string:aid>")
def ack_one(aid: str):
    for a in alarm_history:
        if a["id"] == aid:
            a["acked"]    = True
            a["ts_acked"] = datetime.now().isoformat(timespec="seconds")
            return jsonify({"success": True})
    return jsonify({"success": False, "error": "Not found"}), 404


@bp.post("/clear")
def clear_history():
    alarm_history.clear()
    return jsonify({"success": True})


@bp.get("/export")
def export_excel():
    """Export alarm history to Excel."""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb  = openpyxl.Workbook()
        ws  = wb.active
        ws.title = "Alarm History"
        ws.sheet_view.showGridLines = False

        BG0    = PatternFill("solid", fgColor="04070B")
        BG1    = PatternFill("solid", fgColor="0C1520")
        BG_ACK = PatternFill("solid", fgColor="051505")
        BG_ALM = PatternFill("solid", fgColor="1A0505")
        F_HDR  = Font(bold=True, color="00BCD4", name="Calibri", size=10)
        F_OK   = Font(color="00E676", name="Calibri", size=9)
        F_ALM  = Font(color="FF1744", name="Calibri", size=9)
        F_WARN = Font(color="FF9100", name="Calibri", size=9)
        BORDER = Border(
            left=Side(style="thin", color="1A2D42"), right=Side(style="thin", color="1A2D42"),
            top=Side(style="thin", color="1A2D42"),  bottom=Side(style="thin", color="1A2D42"),
        )
        CTR = Alignment(horizontal="center", vertical="center")
        LFT = Alignment(horizontal="left",   vertical="center")

        # Title
        ws.merge_cells("A1:G1")
        c = ws["A1"]; c.value = "PLC ALARM HISTORY REPORT"; c.font = Font(bold=True, color="FF1744", name="Calibri", size=14)
        c.fill = BG0; c.alignment = CTR; ws.row_dimensions[1].height = 28

        # Subtitle
        ws.merge_cells("A2:G2")
        c = ws["A2"]; c.value = f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} · Total: {len(alarm_history)} · Unacked: {sum(1 for x in alarm_history if not x['acked'])}"
        c.font = Font(color="4A6880", name="Calibri", size=9); c.fill = BG0; c.alignment = CTR

        # Headers
        headers    = ["ID", "Raised", "Acknowledged At", "PLC", "Device", "Message", "Severity"]
        col_widths = [10,   20,        20,                16,    12,       40,          12]
        for ci, (h, w) in enumerate(zip(headers, col_widths), 1):
            c = ws.cell(row=3, column=ci, value=h)
            c.font = F_HDR; c.fill = BG0; c.border = BORDER; c.alignment = CTR
            ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = w
        ws.row_dimensions[3].height = 18

        for ri, a in enumerate(alarm_history, 4):
            row_fill = BG_ACK if a["acked"] else BG_ALM
            row_font = F_OK   if a["acked"] else (F_ALM if a["severity"] == "fault" else F_WARN)
            vals = [a["id"], a["ts_raised"], a.get("ts_acked") or "—",
                    a["plc_name"], a["device"], a["message"], a["severity"].upper()]
            for ci, v in enumerate(vals, 1):
                cc = ws.cell(row=ri, column=ci, value=v)
                cc.fill = row_fill; cc.border = BORDER; cc.font = row_font
                cc.alignment = CTR if ci != 6 else LFT
            ws.row_dimensions[ri].height = 13

        fname = f"AlarmHistory_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        fpath = os.path.join(_LOGS_DIR, fname)
        wb.save(fpath)
        return send_from_directory(_LOGS_DIR, fname, as_attachment=True)
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500
