"""
API — Full PLC Backup (GX Works-style)
========================================
Performs a complete device memory backup similar to GX Works
"Read from PLC" — Device Memory tab.

What is backed up (all ranges read in chunks):
  Word devices : D (Data), TN (Timer current), CN (Counter current),
                 SD (Special data), W (Link registers)
  Bit  devices : M (Internal relays), L (Latch relays),
                 F (Annunciators), SM (Special relays)

Output : JSON  — full machine-readable memory dump
         Excel — formatted report (non-zero values highlighted)

Routes:
  POST /api/backup/<plc_id>             → run full backup
  GET  /api/backup/list                 → list saved backups
  GET  /api/backup/download/<fname>     → download JSON backup
  GET  /api/backup/excel/<fname>        → download Excel report
  POST /api/backup/restore/<plc_id>    → write backup data back to PLC
"""

import json
import os
import time
from datetime import datetime

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from flask import Blueprint, jsonify, request, send_from_directory
from app.config import PLC_CONFIG
from app.plc.poller import connections

bp = Blueprint("backup", __name__, url_prefix="/api/backup")

# ── Paths ──────────────────────────────────────────────────────────────────
_ROOT        = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
_BACKUPS_DIR = os.path.join(_ROOT, "backups")
os.makedirs(_BACKUPS_DIR, exist_ok=True)

# ── Read chunk limits (safe for Type 3E binary mode) ──────────────────────
_WORD_CHUNK = 900    # Type 3E max = 960 words
_BIT_CHUNK  = 3600   # Type 3E max = 7168 bits

# ── Device ranges to back up (mirrors GX Works device memory) ─────────────
# (device_prefix, start_addr, count, type, description)
BACKUP_PLAN = [
    # Word devices
    ("D",  0,     8000, "word", "Data Registers"),
    ("TN", 0,     512,  "word", "Timer Current Values"),
    ("CN", 0,     256,  "word", "Counter Current Values"),
    ("SD", 0,     512,  "word", "Special Data Registers"),
    ("W",  0,     512,  "word", "Link Registers"),
    # Bit devices
    ("M",  0,     8192, "bit",  "Internal Relays"),
    ("L",  0,     8192, "bit",  "Latch Relays"),
    ("F",  0,     256,  "bit",  "Annunciators"),
    ("SM", 0,     512,  "bit",  "Special Auxiliary Relays"),
]


# ── Low-level chunked readers ──────────────────────────────────────────────

def _read_all_words(conn, prefix: str, start: int, count: int) -> list:
    """Read word devices in chunks. Returns list of ints (0 on read error)."""
    data = []
    pos  = start
    while pos < start + count:
        n    = min(_WORD_CHUNK, (start + count) - pos)
        head = f"{prefix}{pos}"
        vals = conn.read_words(head, n)
        data.extend(vals if vals is not None else [0] * n)
        pos += n
    return data


def _read_all_bits(conn, prefix: str, start: int, count: int) -> list:
    """Read bit devices in chunks. Returns list of 0/1 (0 on read error)."""
    data = []
    pos  = start
    while pos < start + count:
        n    = min(_BIT_CHUNK, (start + count) - pos)
        head = f"{prefix}{pos}"
        vals = conn.read_bits(head, n)
        data.extend(vals if vals is not None else [0] * n)
        pos += n
    return data


# ── Excel report generator ─────────────────────────────────────────────────

def _make_excel_report(backup: dict, path: str):
    """Generate a styled Excel workbook from the backup data."""
    wb  = openpyxl.Workbook()
    BG0 = PatternFill("solid", fgColor="04070B")
    BG1 = PatternFill("solid", fgColor="0C1520")
    BG2 = PatternFill("solid", fgColor="0A1018")
    BG_NZ = PatternFill("solid", fgColor="0D1F0D")   # non-zero rows
    BG_HDR = PatternFill("solid", fgColor="061018")

    F_TITLE = Font(bold=True, color="00BCD4", name="Calibri", size=14)
    F_HDR   = Font(bold=True, color="00BCD4", name="Calibri", size=10)
    F_NZ    = Font(color="00E676",            name="Calibri", size=10)
    F_ZERO  = Font(color="2A4055",            name="Calibri", size=10)
    F_WARN  = Font(bold=True, color="FF9100", name="Calibri", size=10)
    F_SUB   = Font(color="4A6880",            name="Calibri", size=9)
    F_META  = Font(color="C8DDF0",            name="Calibri", size=10)

    BORDER = Border(
        left=Side(style="thin", color="1A2D42"),
        right=Side(style="thin", color="1A2D42"),
        top=Side(style="thin", color="1A2D42"),
        bottom=Side(style="thin", color="1A2D42"),
    )
    CTR = Alignment(horizontal="center", vertical="center")
    LFT = Alignment(horizontal="left",   vertical="center")

    meta  = backup["metadata"]
    stats = backup["summary"]

    # ── Cover sheet ────────────────────────────────────────────────────────
    ws = wb.active
    ws.title = "Cover"
    ws.sheet_view.showGridLines = False
    ws.column_dimensions["A"].width = 26
    ws.column_dimensions["B"].width = 40

    ws.merge_cells("A1:B1")
    c = ws["A1"]; c.value = "PLC DEVICE MEMORY BACKUP"; c.font = F_TITLE
    c.fill = BG0; c.alignment = CTR; ws.row_dimensions[1].height = 32

    ws.merge_cells("A2:B2")
    c = ws["A2"]; c.value = "GX Works-Compatible Device Memory Dump"
    c.font = F_SUB; c.fill = BG0; c.alignment = CTR; ws.row_dimensions[2].height = 16

    fields = [
        ("PLC Name",      meta["plc_name"]),
        ("IP Address",    meta["ip"]),
        ("CPU Model",     meta["model"]),
        ("Backup Time",   meta["backup_time"]),
        ("Software",      meta["software"]),
        ("Total Points",  stats["total_points"]),
        ("Non-Zero",      stats["non_zero"]),
        ("Zero Points",   stats["total_points"] - stats["non_zero"]),
    ]
    for i, (k, v) in enumerate(fields, start=4):
        c1 = ws.cell(row=i, column=1, value=k)
        c2 = ws.cell(row=i, column=2, value=str(v))
        c1.font = F_HDR;  c1.fill = BG1; c1.border = BORDER; c1.alignment = LFT
        c2.font = F_META; c2.fill = BG2; c2.border = BORDER; c2.alignment = LFT
        if k == "Non-Zero" and stats["non_zero"] > 0:
            c2.font = Font(color="00E676", bold=True, name="Calibri", size=10)

    # Device summary
    row = len(fields) + 6
    ws.cell(row=row, column=1, value="DEVICE").font = F_HDR
    ws.cell(row=row, column=2, value="NON-ZERO / TOTAL").font = F_HDR
    for cell in [ws.cell(row=row, column=1), ws.cell(row=row, column=2)]:
        cell.fill = BG_HDR; cell.border = BORDER; cell.alignment = CTR
    row += 1
    for dev, info in stats["by_device"].items():
        c1 = ws.cell(row=row, column=1, value=dev)
        c2 = ws.cell(row=row, column=2, value=f"{info['non_zero']} / {info['count']}")
        c1.font = Font(color="00BCD4", name="Calibri", size=10)
        c2.font = F_NZ if info["non_zero"] else F_ZERO
        c1.fill = BG1; c2.fill = BG1
        c1.border = BORDER; c2.border = BORDER
        c1.alignment = CTR; c2.alignment = CTR
        row += 1

    # ── One sheet per device ───────────────────────────────────────────────
    for dev, dev_info in backup["devices"].items():
        dtype  = dev_info["type"]
        data   = dev_info["data"]
        label  = dev_info["label"]
        start  = dev_info["start"]

        ws2 = wb.create_sheet(dev)
        ws2.sheet_view.showGridLines = False
        ws2.freeze_panes = "A4"

        ws2.merge_cells("A1:D1")
        c = ws2["A1"]
        c.value = f"{dev} — {label}  ({len(data)} points)"
        c.font = F_TITLE; c.fill = BG0; c.alignment = CTR
        ws2.row_dimensions[1].height = 26

        hdrs = ["Address", "Decimal", "Hex", "Binary" if dtype == "bit" else "Hex16"]
        col_widths = [14, 12, 10, 18]
        for ci, (h, w) in enumerate(zip(hdrs, col_widths), 1):
            c = ws2.cell(row=3, column=ci, value=h)
            c.font = F_HDR; c.fill = BG_HDR; c.border = BORDER; c.alignment = CTR
            ws2.column_dimensions[get_column_letter(ci)].width = w
        ws2.row_dimensions[3].height = 18

        row_num = 4
        for idx, val in enumerate(data):
            addr = start + idx
            nz   = val != 0
            fill = BG_NZ if nz else BG1

            addr_str = f"{dev}{addr}"
            if dtype == "bit":
                extra = bin(val)[2:] if nz else "0"
            else:
                extra = f"0x{val:04X}" if nz else "0x0000"

            cells_data = [addr_str, val, f"0x{addr:04X}", extra]
            for ci, cv in enumerate(cells_data, 1):
                cc = ws2.cell(row=row_num, column=ci, value=cv)
                cc.fill   = fill
                cc.border = BORDER
                cc.font   = F_NZ if nz else F_ZERO
                cc.alignment = CTR
            ws2.row_dimensions[row_num].height = 13
            row_num += 1

    wb.save(path)


# ── API Routes ─────────────────────────────────────────────────────────────

@bp.post("/<int:pid>")
def create_backup(pid: int):
    """
    Full device memory backup — reads all configured device ranges
    in chunks and saves JSON + Excel to disk.
    """
    cfg = next((c for c in PLC_CONFIG if c["id"] == pid), None)
    if not cfg:
        return jsonify({"success": False, "error": "PLC not found"}), 404

    conn = connections.get(pid)
    if not conn or not conn.connected:
        # Try to connect
        if conn:
            conn._connect()
        if not conn or not conn.connected:
            return jsonify({"success": False, "error": f"Cannot connect to {cfg['ip']}"}), 503

    ts        = datetime.now()
    ts_str    = ts.strftime("%Y%m%d_%H%M%S")
    base_name = f"backup_{cfg['name']}_{ts_str}"
    json_file = base_name + ".json"
    xlsx_file = base_name + ".xlsx"
    t0        = time.time()

    devices    = {}
    by_device  = {}
    total_pts  = 0
    total_nz   = 0

    for (prefix, start, count, dtype, label) in BACKUP_PLAN:
        if dtype == "word":
            data = _read_all_words(conn, prefix, start, count)
        else:
            data = _read_all_bits(conn, prefix, start, count)

        nz_count = sum(1 for v in data if v != 0)
        devices[prefix] = {
            "label":  label,
            "type":   dtype,
            "start":  start,
            "count":  count,
            "data":   data,
        }
        by_device[prefix] = {"count": count, "non_zero": nz_count}
        total_pts += count
        total_nz  += nz_count

    elapsed_ms = round((time.time() - t0) * 1000)

    backup_doc = {
        "metadata": {
            "plc_name":    cfg["name"],
            "ip":          cfg["ip"],
            "model":       cfg["model"],
            "location":    cfg["location"],
            "backup_time": ts.isoformat(),
            "software":    "PLC Network Monitor v1 (MC Protocol Type 3E)",
            "version":     "3.0",
        },
        "summary": {
            "total_points": total_pts,
            "non_zero":     total_nz,
            "duration_ms":  elapsed_ms,
            "by_device":    by_device,
        },
        "devices": devices,
    }

    # Save JSON
    json_path = os.path.join(_BACKUPS_DIR, json_file)
    with open(json_path, "w") as f:
        json.dump(backup_doc, f, indent=2)

    # Generate Excel report
    xlsx_path = os.path.join(_BACKUPS_DIR, xlsx_file)
    try:
        _make_excel_report(backup_doc, xlsx_path)
        xlsx_size = round(os.path.getsize(xlsx_path) / 1024, 1)
    except Exception as exc:
        xlsx_file = None
        xlsx_size = 0
        print(f"Excel generation error: {exc}")

    json_size = round(os.path.getsize(json_path) / 1024, 1)

    return jsonify({
        "success":     True,
        "json_file":   json_file,
        "excel_file":  xlsx_file,
        "json_size":   f"{json_size} KB",
        "excel_size":  f"{xlsx_size} KB",
        "duration_ms": elapsed_ms,
        "summary": {
            "total_points": total_pts,
            "non_zero":     total_nz,
            "by_device":    by_device,
        },
    })


@bp.get("/list")
def list_backups():
    """Return a list of saved backup files."""
    files = []
    try:
        for fname in sorted(os.listdir(_BACKUPS_DIR), reverse=True):
            fpath = os.path.join(_BACKUPS_DIR, fname)
            if not os.path.isfile(fpath):
                continue
            size_kb = round(os.path.getsize(fpath) / 1024, 1)
            mtime   = datetime.fromtimestamp(os.path.getmtime(fpath)).isoformat()
            files.append({
                "filename":  fname,
                "size":      f"{size_kb} KB",
                "modified":  mtime,
                "type":      "excel" if fname.endswith(".xlsx") else "json",
            })
    except Exception as exc:
        return jsonify({"success": False, "error": str(exc)}), 500

    return jsonify({"success": True, "files": files})


@bp.get("/download/<string:fname>")
def download_backup(fname: str):
    """Download a backup JSON file."""
    return send_from_directory(_BACKUPS_DIR, fname, as_attachment=True)


@bp.get("/excel/<string:fname>")
def download_excel(fname: str):
    """Download a backup Excel report."""
    return send_from_directory(_BACKUPS_DIR, fname, as_attachment=True)


@bp.post("/restore/<int:pid>")
def restore_backup(pid: int):
    """
    Restore device memory from a JSON backup file.
    Body: { "filename": "backup_PLC-LINE-01_20260319_232704.json",
            "devices": ["D", "M"]  ← optional: restore only these devices }
    """
    cfg = next((c for c in PLC_CONFIG if c["id"] == pid), None)
    if not cfg:
        return jsonify({"success": False, "error": "PLC not found"}), 404

    body     = request.json or {}
    filename = body.get("filename")
    selected = body.get("devices", None)   # None = restore all

    if not filename:
        return jsonify({"success": False, "error": "filename required"}), 400

    path = os.path.join(_BACKUPS_DIR, filename)
    if not os.path.exists(path):
        return jsonify({"success": False, "error": "File not found"}), 404

    conn = connections.get(pid)
    if not conn or not conn.connected:
        if conn:
            conn._connect()
        if not conn or not conn.connected:
            return jsonify({"success": False, "error": f"Cannot connect to {cfg['ip']}"}), 503

    with open(path) as f:
        backup_doc = json.load(f)

    restored = {}
    errors   = {}

    for dev, info in backup_doc["devices"].items():
        if selected and dev not in selected:
            continue

        data  = info["data"]
        dtype = info["type"]
        start = info["start"]

        # Write in chunks
        pos = start
        ok  = True
        while pos < start + len(data):
            chunk = data[pos - start : pos - start + _WORD_CHUNK]
            head  = f"{dev}{pos}"
            if dtype == "word":
                result = conn.write_words(head, chunk)
            else:
                result = conn.write_bits(head, chunk)
            if not result:
                ok = False
                errors[dev] = conn.last_err
                break
            pos += len(chunk)

        restored[dev] = "ok" if ok else f"error: {errors.get(dev)}"

    return jsonify({
        "success":  len(errors) == 0,
        "restored": restored,
        "errors":   errors,
    })
